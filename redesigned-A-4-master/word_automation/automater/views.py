

import datetime
import random
import pythoncom
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render,redirect
import pandas as pd
from.models import excel
from docxtpl import DocxTemplate
from docx.shared import Cm,Inches,Mm
from docxtpl import DocxTemplate,InlineImage
import pandas as pd
import random
import numpy as np
import jinja2
import sys
import math
import os
import comtypes.client
import datetime
from openpyxl import Workbook
from io import StringIO
import zipfile
import os, io
import math
import glob
from docx2pdf import convert

# Create your views here.
def home(request):
    pythoncom.CoInitialize()


    if request.method=='POST' and request.FILES:
        file=request.FILES['excel_file']
        obj=excel.objects.create(excel_file=file)
        name=str(file)
        l=name.split('.')


        if l[1]=='xlsx' or l[1]=='xls':


            #function to get random values of criteria
            def ra(value):
                if '±' in value:
                    p=value.split('±')
                    for i in range(len(p)):
                        p[i]=float(p[i])
                    x=round(random.uniform(p[0]-p[1],p[0]+p[1]), 2)
                    return x
                elif '+' in value:
                    p=value.split('+')
                    for i in range(len(p)):
                        p[i]=float(p[i])
                    x=round(random.uniform(p[0],p[0]+p[1]),2)
                    return x 
                elif '-' in value:
                    p=value.split('-')
                    for i in range(len(p)):
                        p[i]=float(p[i])
                    x=round(random.uniform(p[0]-p[1],p[0]),2)
                    return x 
                else:
                   
                    return value





            # doc=DocxTemplate('files/COAtemplate1.docx')
            df = pd.read_excel(file)
            path_df=pd.read_excel(file,sheet_name='path')
            pathid=path_df['pathid'][0]
            req_path_temp=f'files/{pathid}'
            df1=pd.read_excel(f'{req_path_temp}/data.xlsx')
            ProductCode=df1['ProductCode'].values
            DeviceName=df1['DeviceName'].values
            expiry=df1['EXP'].values
            description=df1['Description'].values
            ifu=df1['IFU'].values
            iso=df1['ISO'].values
            sam_plan=df1['sampling plan'].values
            StorageCondition=df1['StorageCondition'].values
            para1=df1['parameter1'].values
            para2=df1['parameter2'].values
            para3=df1['parameter3'].values
            para4=df1['parameter4'].values
            para5=df1['parameter5'].values
            para6=df1['parameter6'].values
            para7=df1['parameter7'].values
            para8=df1['parameter8'].values
            para9=df1['parameter9'].values
            para10=df1['parameter10'].values
            criteria1=df1['criteria1'].values
            criteria2=df1['criteria2'].values
            criteria3=df1['criteria3'].values
            criteria4=df1['criteria4'].values
            criteria5=df1['criteria5'].values
            criteria6=df1['criteria6'].values
            criteria7=df1['criteria7'].values
            criteria8=df1['criteria8'].values
            criteria9=df1['criteria9'].values
            criteria10=df1['criteria10'].values
            inspec1=df1['inspectionMode1'].values
            inspec2=df1['inspectionMode2'].values
            inspec3=df1['inspectionMode3'].values
            inspec4=df1['inspectionMode4'].values
            inspec5=df1['inspectionMode5'].values
            inspec6=df1['inspectionMode6'].values
            inspec7=df1['inspectionMode7'].values
            inspec8=df1['inspectionMode8'].values
            inspec9=df1['inspectionMode9'].values
            inspec10=df1['inspectionMode10'].values



            context={}

            for i in range(len(ProductCode)):
                

    
                case={
            ProductCode[i]:{
            'ProductCode':ProductCode[i],
            'DeviceName':DeviceName[i],
            'expiry':expiry[i],
            'description':description[i],
            'ifu':ifu[i],
            'iso':iso[i],
            'sampling plan':sam_plan[i],
            'StorageCondition':StorageCondition[i],
            'para1':para1[i],
            'para2':para2[i],
            'para3':para3[i],
            'para4':para4[i],
            'para5':para5[i],
            'para6':para6[i],
            'para7':para7[i],
            'para8':para8[i],
            'para9':para9[i],
            'para10':para10[i],
            'criteria1':inspec1[i],
            'criteria2':inspec2[i],
            'criteria3':inspec3[i],
            'criteria4':inspec4[i],
            'criteria5':inspec5[i],
            'criteria6':inspec6[i],
            'criteria7':inspec7[i],
            'criteria8':inspec8[i],
            'criteria9':inspec9[i],
            'criteria10':inspec10[i],
            'inspec1':criteria1[i],
            'inspec2':criteria2[i],
            'inspec3':criteria3[i],
            'inspec4':criteria4[i],
            'inspec5':criteria5[i],
            'inspec6':criteria6[i],
            'inspec7':criteria7[i],
            'inspec8':criteria8[i],
            'inspec9':criteria9[i],
            'inspec10':criteria10[i],
            }
        }
    #a way to access the context2 with name of material in context
                context[ProductCode[i]]=case
      
            context2={}
            c1=df['ProductCode'].values
            c2 = df['LotNo'].values
            c3 = df['LotSize'].values
            c4 = df['COAdate'].values
            l1=[]
            l2=[]
            book = openpyxl.load_workbook(file)
            sheet= book.get_sheet_by_name('Sheet1')
            sheet['G1']='COAStatus'
            sheet['H1']='DOCStatus'
            for i in range(len(c4)):
                test_date=pd.Timestamp(np.datetime64(c4[i])).to_pydatetime()
                test_date=str(test_date.date())
                test_date=datetime.datetime.strptime(test_date, "%Y-%m-%d").strftime("%d-%m-%Y")
                l1.append(test_date)
            c4=l1
            c5 = df['MFGdate'].values
            for i in range(len(c5)):
                test_date=pd.Timestamp(np.datetime64(c5[i])).to_pydatetime()
                test_date=str(test_date.date())
                test_date=datetime.datetime.strptime(test_date, "%Y-%m-%d").strftime("%m/%Y").replace('/','-')
                slice=test_date[3:5]
                test_date=test_date.replace(slice,"")
                l2.append(test_date)
            c5=l2
          

     
            c7 = df['InspectedBy'].values  
            df3=pd.read_excel(f'{req_path_temp}/data.xlsx',sheet_name='admin')
            coa_admin=df3['COA'].values
            doc_admin=df3['DOC'].values
            df4=pd.read_excel(file,sheet_name='doc')
            check_coa=df4['downloaddoc'].values[0]
            check_doc=df4['downloaddoc'].values[1]
            check_pdf_coa=df4['pdf'].values[0]
            check_pdf_doc=df4['pdf'].values[1]
            admin_context={}
            flag=1
            for i in range(len(c1)):

                procode=c1[i]
                admin_context[procode]={
                    'COA':check_coa,
                    'DOC':check_doc,
                }
              
                if (procode in context) and (coa_admin[i]==1):
                    
               
                    sheet.cell(row=i+2, column=7).value = 'ok'
                else:
                   
                    sheet.cell(row=i+2, column=7).value = 'ERROR'
                    continue
                if (procode in context) and (doc_admin[i]==1):
                    
               
                    sheet.cell(row=i+2, column=8).value = 'ok'
                else:
                   
                    sheet.cell(row=i+2, column=8).value = 'ERROR'
                    continue
                    
                if os.path.exists(f"{req_path_temp}/sign/{c7[i]}.png"):
                    img_d=f"{req_path_temp}/sign/{c7[i]}.png"
                else:
                    img_d=f"{req_path_temp}/sign/random.png"

                
                

                dd=c5[i].split('-')
                dd[0]=int(dd[0])
                dd[1]=int(dd[1])
                expp=context[c1[i]][c1[i]]['expiry']
          
                if math.isnan(expp):
                 
                    exp_date='NA'
                
                else:
                    expp=float(expp)
                    exp_y=math.floor(expp)
                    exp_m=math.floor((expp-exp_y)*12)
                    dd[0]=dd[0]+exp_m
                    dd[1]=dd[1]+exp_y 
                    dd[0]=str(dd[0])
                    dd[1]=str(dd[1])
                    exp_date=dd[0]+'-'+dd[1]
               
                    
                    



                lotsize=c3[i]

                sampling_plan=context[c1[i]][c1[i]]['sampling plan']
                
                df2=pd.read_excel(f'{req_path_temp}/data.xlsx',sheet_name='sampling plan')
                mini=df2['min'].values 
                maxi=df2['max'].values
             
                cc=0
                for j in range(len(mini)):
                    cc+=1
                 
                    if (lotsize>=mini[j]) and (lotsize<maxi[j]):
                        req=df2[sampling_plan].values[cc-1]
                        break 
                cc=0
               
               
                



             
                req_l=[]
                p=[]
                prev=[]
                
                avg1=avg2=avg3=avg4=avg5=avg6=avg7=avg8=avg9=avg10=0
                list_of_inspec=[context[c1[i]][c1[i]]['inspec1'],context[c1[i]][c1[i]]['inspec2'],context[c1[i]][c1[i]]['inspec3'],context[c1[i]][c1[i]]['inspec4'],context[c1[i]][c1[i]]['inspec5'],context[c1[i]][c1[i]]['inspec6'],context[c1[i]][c1[i]]['inspec7'],context[c1[i]][c1[i]]['inspec8'],context[c1[i]][c1[i]]['inspec9'],context[c1[i]][c1[i]]['inspec10']]
                for m in range(int(req)):
                    for val in list_of_inspec:
                        v=ra(val)
                        p.append(v)
                    prev.append(p)
                    
                    req_l.append({'label':m+1,'cols':p})
                    if isinstance(p[0], str):
                        avg1='C'
                    else:
                        avg1+=p[0]
                    if isinstance(p[1], str):
                        avg2='C'
                    else:
                        avg2+=p[1]
                    if isinstance(p[2], str):
                        avg3='C'
                    else:
                        avg3+=p[2]
                    if isinstance(p[3], str):
                        avg4='C'
                    else:
                        avg4+=p[3]
                    if isinstance(p[4], str):
                        avg5='C'
                    else:
                        avg5+=p[4]
                    if isinstance(p[5], str):
                        avg6='C'
                    else:
                        avg6+=p[5]
                    if isinstance(p[6], str):
                        avg7='C'
                    else:
                        avg7+=p[6]
                    if isinstance(p[7], str):
                        avg8='C'
                    else:
                        avg8+=p[7]
                    if isinstance(p[8], str):
                        avg9='C'
                    else:
                        avg9+=p[8]
                    if isinstance(p[9], str):
                        avg10='C'
                    else:
                        avg10+=p[9]

                    p=[]
                
        
                try:
                    avg1=avg1/req 
                    avg1= round(avg1, 2)
                except:
                    avg1=avg1
                try:
                    avg2=avg2/req
                    avg2= round(avg2, 2) 
                except:
                    avg2=avg2
                try:
                    avg3=avg3/req 
                    avg3= round(avg3, 2)
                except:
                    avg3=avg3
                try:
                    avg4=avg4/req
                    avg4= round(avg4, 2) 
                except:
                    avg4=avg4
                try:
                    avg5=avg5/req 
                    avg5= round(avg5, 2)
                except:
                    avg5=avg5
                try:
                    avg6=avg6/req 
                    avg6= round(avg6, 2)
                except:
                    avg6=avg6
                try:
                    avg7=avg7/req 
                    avg7= round(avg7, 2)
                except:
                    avg7=avg7
                try:
                    avg8=avg8/req 
                    avg8= round(avg8, 2)
                except:
                    avg8=avg8
                try:
                    avg9=avg9/req 
                    avg9= round(avg9, 2)
                except:
                    avg9=avg9
                try:
                    avg10=avg10/req
                    avg10= round(avg10, 2) 
                except:
                    avg10=avg10
                v=0
                main_sd1=main_sd2=main_sd3=main_sd4=main_sd5=main_sd6=main_sd7=main_sd8=main_sd9=main_sd10=0
                for sd in range(int(req)):
                    for mom in range(len(prev[0])):
                        
                        if v==0:
                                
                            main_sd1=main_sd1+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                              
                        if v==1:
                            try:
                                main_sd2=main_sd2+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd2=''
                        if v==2:
                            try:
                                main_sd3=main_sd3+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd3=''
                         
                        if v==3:
                            try:
                                main_sd4=main_sd4+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd4=''
                        
                        if v==4:
                            try:
                                main_sd5=main_sd5+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd5=''
                            
                        if v==5:
                            try:
                                main_sd6=main_sd6+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd6=''
                         
                        if v==6:
                            try:
                                main_sd7=main_sd7+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd7=''
                               
                        if v==7:
                            try:
                                main_sd8=main_sd8+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd8=''
                               
                        if v==8:
                            try:
                                main_sd9=main_sd9+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd9=''
                          
                        if v==9:
                            try:
                                main_sd10=main_sd10+math.sqrt(math.pow(abs(prev[v][mom]-avg1),2)/9)
                            except:
                                main_sd10=''
                        v+=1
                try:
                    main_sd1=round(main_sd1,3)
                except:
                    main_sd1=main_sd1
                try:
                    main_sd2=round(main_sd2,3)
                except:
                    main_sd2=main_sd2
                try:
                    main_sd3=round(main_sd3,3)
                except:
                    main_sd3=main_sd3
                try:
                    main_sd4=round(main_sd4,3)
                except:
                    main_sd4
                try:
                    main_sd5=round(main_sd5,3)
                except:
                    main_sd5
                try:
                    main_sd6=round(main_sd6,3)
                except:
                    main_sd6=main_sd6
                try:
                    main_sd7=round(main_sd7,3)
                except:
                    main_sd7=main_sd7
                try:
                    main_sd8=round(main_sd8,3)
                except:
                    main_sd8=main_sd8
                try:
                    main_sd9=round(main_sd9,3)
                except:
                    main_sd9=main_sd9
                try:
                    main_sd10=round(main_sd10,3)
                except:
                    main_sd10=main_sd10
                              
                        
                
             
                context2={
                    'DeviceName':context[c1[i]][c1[i]]['DeviceName'],
                    'BatchNo':c2[i],
                    'Description':context[c1[i]][c1[i]]['description'],
                    'IFU':context[c1[i]][c1[i]]['ifu'],
                    'ISO':context[c1[i]][c1[i]]['iso'],
                    'InspectedBy':c7[i],
                    'COAdate':c4[i],
                    'LotNo':c2[i],
                    'LotSize':c3[i],
                    'MFGdate':c5[i],
                    'EXPdate':exp_date,
                    'ProductCode':c1[i],
                    'SampleSize':int(req),
                    'parameter1':context[c1[i]][c1[i]]['para1'],
                    'parameter2':context[c1[i]][c1[i]]['para2'],
                    'parameter3':context[c1[i]][c1[i]]['para3'],
                    'parameter4':context[c1[i]][c1[i]]['para4'],
                    'parameter5':context[c1[i]][c1[i]]['para5'],
                    'parameter6':context[c1[i]][c1[i]]['para6'],
                    'parameter7':context[c1[i]][c1[i]]['para7'],
                    'parameter8':context[c1[i]][c1[i]]['para8'],
                    'parameter9':context[c1[i]][c1[i]]['para9'],
                    'parameter10':context[c1[i]][c1[i]]['para10'],
                    'criteria1':context[c1[i]][c1[i]]['inspec1'],
                    'criteria2':context[c1[i]][c1[i]]['inspec2'],
                    'criteria3':context[c1[i]][c1[i]]['inspec3'],
                    'criteria4':context[c1[i]][c1[i]]['inspec4'],
                    'criteria5':context[c1[i]][c1[i]]['inspec5'],
                    'criteria6':context[c1[i]][c1[i]]['inspec6'],
                    'criteria7':context[c1[i]][c1[i]]['inspec7'],
                    'criteria8':context[c1[i]][c1[i]]['inspec8'],
                    'criteria9':context[c1[i]][c1[i]]['inspec9'],
                    'criteria10':context[c1[i]][c1[i]]['inspec10'],
                    'inspectionMode1':context[c1[i]][c1[i]]['criteria1'],
                    'inspectionMode2':context[c1[i]][c1[i]]['criteria2'],
                    'inspectionMode3':context[c1[i]][c1[i]]['criteria3'],
                    'inspectionMode4':context[c1[i]][c1[i]]['criteria4'],
                    'inspectionMode5':context[c1[i]][c1[i]]['criteria5'],
                    'inspectionMode6':context[c1[i]][c1[i]]['criteria6'],
                    'inspectionMode7':context[c1[i]][c1[i]]['criteria7'],
                    'inspectionMode8':context[c1[i]][c1[i]]['criteria8'],
                    'inspectionMode9':context[c1[i]][c1[i]]['criteria9'],
                    'inspectionMode10':context[c1[i]][c1[i]]['criteria10'],
                    'tbl_contents':req_l,
                    'avg1':avg1,
                    'avg2':avg2,
                    'avg3':avg3,
                    'avg4':avg4,
                    'avg5':avg5,
                    'avg6':avg6,
                    'avg7':avg7,
                    'avg8':avg8,
                    'avg9':avg9,
                    'avg10':avg10,
                    'sd1':main_sd1,
                    'sd2':main_sd2,
                    'sd3':main_sd3,
                    'sd4':main_sd4,
                    'sd5':main_sd5,
                    'sd6':main_sd6,
                    'sd7':main_sd7,
                    'sd8':main_sd8,
                    'sd9':main_sd9,
                    'sd10':main_sd10,
                    'win':'C',


                }
            
                list_of_docs=[]
                if admin_context[c1[i]]['COA']==1:
                    list_of_docs.append(f'{req_path_temp}/templates/COA.docx')
                if admin_context[c1[i]]['DOC']==1:
                    list_of_docs.append(f'{req_path_temp}/templates/DOC.docx')
                
                for docs in list_of_docs:
                    doc=DocxTemplate(docs)
                    obj1=InlineImage(doc,image_descriptor=img_d,width=Mm(20), height=Mm(10))
                    context2['sign']=obj1
                    doc.render(context2)
                    if docs.split('.')[0][-1:-4:-1][::-1]=='COA':
                        doc.save(f'output/COAfiles/COA {c1[i]} {c2[i]}.docx')
                        if check_pdf_coa:
                            convert(f'output/COAfiles/COA {c1[i]} {c2[i]}.docx',f'output/COAfiles/COA {c1[i]} {c2[i]}.pdf')
                            os.remove(f'output/COAfiles/COA {c1[i]} {c2[i]}.docx')
                        
                    else:
                        doc.save(f'output/DOCfiles/DOC {c1[i]} {c2[i]}.docx')
                        if check_pdf_doc:
                            convert(f'output/DOCfiles/DOC {c1[i]} {c2[i]}.docx',f'output/DOCfiles/DOC {c1[i]} {c2[i]}.pdf')
                            os.remove(f'output/DOCfiles/DOC {c1[i]} {c2[i]}.docx')


               
                 
            book.save('output/COAStatusReport.xlsx')
            





        else:
            return HttpResponse("Please enter an excel file only")
        return render(request,'automater/home.html',{'file':obj})
    
    return render(request,'automater/home.html')
def download(request):
    l=glob.glob('output/COAfiles/*.*')
    filenames = l
    filenames.append('output/COAStatusReport.xlsx')
    docs=glob.glob('output/DOCfiles/*.*')
    for doc in docs:
        filenames.append(doc)
    
   


    zip_subdir = "COA Output"
    zip_filename = "%s.zip" % zip_subdir

 
    s = io.BytesIO()

    zf = zipfile.ZipFile(s, "w")
    for fpath in filenames:
        fdir, fname = os.path.split(fpath)
        zip_path = os.path.join('', fname)
        if zip_path.split('.')[1]=='docx' or zip_path.split('.')[1]=='pdf':
            if fname[0:3]=='COA':
                zf.write(fpath, f'/COAfiles/{zip_path}')
            else:
                zf.write(fpath,f'/DOCfiles/{zip_path}')
                

        else:
            zf.write(fpath, f'{zip_path}')
    zf.close()
    resp = HttpResponse(s.getvalue(), content_type = "application/x-zip-compressed")
    resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename
    return resp
def refresh(request):
    try:
        rem_xl='output/COAStatusReport.xlsx'
        os.remove(rem_xl)
        #removing all the output files in the zip folder after user downloads them (caching)
        removefiles=glob.glob('output/COAfiles/*.*')
        for i in removefiles:
            rem=i.replace('\\','/')
            os.remove(rem)
        removefiles1=glob.glob('output/DOCfiles/*.*')
        for i in removefiles1:
            rem=i.replace('\\','/')
            os.remove(rem)
        #removing all the input media files after user downloads the output zip folder
        media_files=glob.glob('media/*.xlsx')
        for mfile in media_files:
            mfile=mfile.replace('\\','/')
            os.remove(mfile)
        return redirect('/')
    except:
        return redirect('/')